from openpyxl import Workbook
from openpyxl import load_workbook

def create_Exel():
    name = "data.xlsx"

    wb = Workbook()

    sheet = wb.active


    sheet["A1"] = "Username"
    sheet["A2"] = "Password"
    sheet["B1"] = "Me"
    sheet["B2"] = "132"

    wb.save(filename=name)


def read_data():
    wb = load_workbook()
    data = ""
    return data



if __name__=="__main__":
    create_Exel()


import sqlite3

conn = sqlite3.connect("db1.db")

cursor = conn.cursor()
sql = "SELECT * FROM Customers"

results = cursor.execute(sql)